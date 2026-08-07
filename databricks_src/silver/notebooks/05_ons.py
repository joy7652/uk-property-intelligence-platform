# Databricks notebook source
# MAGIC %md
# MAGIC # Silver: Price Index of Private Rents (ONS)
# MAGIC
# MAGIC Reads the newest Price Index of Private Rents vintage from the Bronze Volume and
# MAGIC overwrites `uk_property_intel.silver.ons_private_rents`.
# MAGIC
# MAGIC ONS republishes the whole series each month and revises the latest months, so a
# MAGIC full overwrite from the newest vintage is the correct write.
# MAGIC
# MAGIC The sheet is converted to CSV on cluster-local disk before Spark reads it.
# MAGIC spark-excel returns the cell's display format when reading as string, which is
# MAGIC one decimal place against the six the file holds, and its typed path cannot carry
# MAGIC the `[x]` and `[z]` markers or support `assert_casts_preserved`. openpyxl in
# MAGIC read-only mode returns the stored value and a real date, so the transform
# MAGIC receives the same all-string frame every other source produces.
# MAGIC
# MAGIC The local path makes this notebook single-node only: on a multi-node cluster the
# MAGIC executors do not share the driver's filesystem and the read fails outright, which
# MAGIC is cheap to notice and cheap to fix.
# MAGIC
# MAGIC What the run measured is written to `uk_property_intel.quality.pipeline_run`
# MAGIC and `pipeline_metric` rather than only printed. `run.step()` wraps whichever
# MAGIC cells can raise, so a failure records why before it stops the notebook.

# COMMAND ----------

import csv
import os
import re
from datetime import date, datetime, timezone
from decimal import Decimal

from openpyxl import load_workbook
from pyspark.sql import functions as F
from pyspark.storagelevel import StorageLevel

from databricks_src.quality.audit.writer import AuditRun
from databricks_src.silver.transforms.ons import (
    COLUMN_MAP,
    MEASURE_COLUMNS,
    SILVER_COLUMNS,
    SOURCE_COLUMNS,
    parse_published_date,
    silver_table_ddl,
    transform_ons,
)

VOLUME_PATH = "/Volumes/uk_property_intel/bronze/ons/private_rent_index"
TARGET_TABLE = "uk_property_intel.silver.ons_private_rents"

STAGE_PATH = "/local_disk0/ons_table1.csv"

COVER_SHEET = "Cover sheet"
SHEET = "Table 1"
HEADER_ROW = 3

# Case-insensitive: the landed name is lowercased relative to the source URL.
VINTAGE_PATTERN = re.compile(
    r"^priceindexofprivaterents-(\d{4})-(\d{2})\.xlsx$", re.IGNORECASE
)

# Local file header of a ZIP archive, which is what an xlsx is. The source URL is
# hand-maintained and opaque, so a stale or mistyped path can land an HTML error page
# under the expected filename.
ZIP_MAGIC_BYTES = b"PK\x03\x04"

# Scientific notation would reach Spark as a string. Floats route through Decimal to
# rule it out, and this asserts the rule held.
SCIENTIFIC = re.compile(r"^-?\d+(\.\d+)?[eE][+-]?\d+$")

# One timestamp for the whole run, stamped on every Silver row and carried on the
# audit row, so the two can be joined.
INGESTION_TS = datetime.now(timezone.utc)

# Set when rebuilding from a Bronze copy kept on purpose. The freshness bound cannot
# tell a deliberate rebuild from a stale release.
SKIP_FRESHNESS = False

# COMMAND ----------

run = AuditRun(source="ons", layer="silver", ingestion_ts=INGESTION_TS)
run.start()
print(f"run {run.run_id}")

# COMMAND ----------

# MAGIC %md
# MAGIC ## Select the vintage

# COMMAND ----------

with run.step():
    # Vintage comes from the filename, not modificationTime: the watermark re-fetches
    # the same release under the same name when the URL is not updated, which moves
    # the timestamp without moving the data.
    entries = dbutils.fs.ls(VOLUME_PATH)  # noqa: F821
    vintages = [
        (VINTAGE_PATTERN.match(entry.name).groups(), entry)
        for entry in entries
        if VINTAGE_PATTERN.match(entry.name)
    ]
    if not vintages:
        raise FileNotFoundError(
            f"No file matching {VINTAGE_PATTERN.pattern} under {VOLUME_PATH}. "
            f"Present: {[entry.name for entry in entries]}"
        )

    label, source_file = max(vintages, key=lambda item: item[0])
    SOURCE_PATH = f"{VOLUME_PATH}/{source_file.name}"

    run.measure("vintages_present", len(vintages))
    run.measure("vintage_label", "-".join(label))
    run.measure("source_files", 1)
    run.measure("source_bytes", source_file.size)

print(f"{len(vintages)} vintage(s) present, reading {source_file.name}")
print(f"{source_file.size / 1024 ** 2:,.1f} MB")

# COMMAND ----------

# MAGIC %md
# MAGIC ## Validate
# MAGIC
# MAGIC The magic bytes are checked before the workbook is opened. Phase 1 established
# MAGIC that a pipeline reporting success confirms bytes moved, not that the right bytes
# MAGIC moved, and this source needs its URL rewritten by hand every month.

# COMMAND ----------

with run.step():
    with open(SOURCE_PATH, "rb") as handle:
        magic = handle.read(len(ZIP_MAGIC_BYTES))
    if magic != ZIP_MAGIC_BYTES:
        raise ValueError(
            f"{SOURCE_PATH} is not an xlsx. Expected {ZIP_MAGIC_BYTES!r}, "
            f"found {magic!r}."
        )
print(f"magic bytes ok: {magic!r}")

# COMMAND ----------

# MAGIC %md
# MAGIC ## Convert
# MAGIC
# MAGIC Values are written as published; meaning is the transform's job. Dates become
# MAGIC ISO, floats route through `Decimal` so nothing arrives in scientific notation,
# MAGIC and the markers pass through as the strings ONS wrote.
# MAGIC
# MAGIC `max_row` and `max_column` come from the sheet's declared dimension rather than
# MAGIC from the row iteration, so they cross-check the conversion against a different
# MAGIC code path. A reader that skips rows is the failure this whole approach exists to
# MAGIC avoid, and it would otherwise leave no trace.

# COMMAND ----------


def serialise(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float):
        # repr gives the shortest round-trip form, which is scientific below 1e-4.
        return f"{Decimal(repr(value)):f}"
    return str(value)


with run.step():
    book = load_workbook(SOURCE_PATH, read_only=True, data_only=True)

    # The artefact signal, and the only one the content carries. The filename
    # records which release the watermark asked for, not which one ONS served, and
    # this source cannot be pattern-matched from its URL.
    cover = book[COVER_SHEET]
    published = next(
        (
            row[0]
            for row in cover.iter_rows(max_col=1, values_only=True)
            if row[0] and "originally published" in str(row[0])
        ),
        None,
    )
    # Recorded rather than only printed, and the absence is recorded as a value:
    # a missing line means ONS changed the cover sheet, which a null could not say.
    run.measure(
        "published_line", published or f"none found on '{COVER_SHEET}'"
    )

    # The date on its own, so it can be compared against the release the filename
    # claims. A wording change costs this signal and not the load, so the parse
    # returns nothing rather than raising, and the line above stays as the record.
    published_on = parse_published_date(published)
    if published_on:
        run.measure("published_date", published_on)

    sheet = book[SHEET]
    declared_rows, declared_columns = sheet.max_row, sheet.max_column
    if declared_columns != len(SOURCE_COLUMNS):
        raise ValueError(
            f"{SHEET} declares {declared_columns} columns, expected "
            f"{len(SOURCE_COLUMNS)}."
        )

    read = written = blank = scientific = 0
    with open(STAGE_PATH, "w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        for row in sheet.iter_rows(min_row=HEADER_ROW, values_only=True):
            read += 1
            if all(value is None for value in row):
                blank += 1
                continue
            fields = [serialise(value) for value in row]
            scientific += sum(1 for field in fields if SCIENTIFIC.match(field))
            writer.writerow(fields)
            written += 1
    book.close()

    expected_read = declared_rows - HEADER_ROW + 1
    if read != expected_read:
        raise ValueError(
            f"{SHEET} yielded {read} rows from row {HEADER_ROW}, but its dimension "
            f"declares {expected_read}. openpyxl dropped rows."
        )
    if scientific:
        raise ValueError(f"{scientific} field(s) written in scientific notation.")


print(published or f"no publication line found on '{COVER_SHEET}'")
if published_on:
    # Printed side by side rather than asserted. The July release being published in
    # July is one observation, not a rule, and the recorded pair is what settles
    # whether a mismatch is worth aborting on.
    print(f"published {published_on}, filename claims release "
          f"{'-'.join(label)}")
else:
    print("no publication date parsed, so the filename cannot be cross-checked")
DATA_ROWS = written - 1
print(f"read={read} written={written} blank={blank} data rows={DATA_ROWS:,}")
print(f"{os.path.getsize(STAGE_PATH) / 1024 ** 2:,.1f} MB at {STAGE_PATH}")

# COMMAND ----------

# MAGIC %md
# MAGIC ## Read
# MAGIC
# MAGIC Types are asserted in the transform, never inferred. `FAILFAST` aborts on a row
# MAGIC whose field count does not match the header rather than padding it with nulls.
# MAGIC `nullValue` is pinned so blank cells land as null whatever the runtime default,
# MAGIC which every unpopulated measure depends on.

# COMMAND ----------

with run.step():
    raw = (
        spark.read.option("header", True)  # noqa: F821
        .option("inferSchema", False)
        .option("quote", '"')
        .option("escape", '"')
        .option("nullValue", "")
        .option("mode", "FAILFAST")
        .csv(f"file://{STAGE_PATH}")
    )

    # Every guard is an action. Without this the file is parsed once per guard.
    raw.persist(StorageLevel.DISK_ONLY)

    source_rows = run.measure("source_rows", raw.count())
    if source_rows != DATA_ROWS:
        raise ValueError(
            f"Spark read {source_rows} rows from the staged CSV, which holds "
            f"{DATA_ROWS}."
        )
print(f"{source_rows:,} source rows, {len(raw.columns)} columns")

# COMMAND ----------

# MAGIC %md
# MAGIC ## Discovery
# MAGIC
# MAGIC The guards in the transform abort the run on a marker outside the positions ONS
# MAGIC publishes it in. Run this against a full release before trusting them, and again
# MAGIC whenever ONS changes the sheet.

# COMMAND ----------

# Every non-numeric value across the 36 measures. [x] is expected; [z] is not, and a
# third marker would abort the run. The frame still carries source headers here, so the
# measures are looked up through the map rather than by their Silver names.
measures = set(MEASURE_COLUMNS)
MEASURE_HEADERS = [
    source for source, target in COLUMN_MAP.items() if target in measures
]

print("non-numeric values in measures")
(
    raw.select(
        F.explode(F.array(*[raw[name] for name in MEASURE_HEADERS])).alias("value")
    )
    .filter(F.col("value").isNotNull() & ~F.col("value").rlike(r"^-?\d+(\.\d+)?$"))
    .groupBy("value")
    .count()
    .orderBy(F.desc("count"))
    .show(truncate=False)
)

# What each geography type carries in the parent column. The transform identifies
# Northern Ireland by code or by this column, so it does not depend on the answer.
print("area code prefix against parent column")
(
    raw.select(
        F.substring(raw["Area code"], 1, 3).alias("prefix"),
        raw["Region or country name"].alias("parent"),
    )
    .distinct()
    .orderBy("prefix", "parent")
    .show(60, truncate=False)
)

# COMMAND ----------

# MAGIC %md
# MAGIC ## Transform
# MAGIC
# MAGIC Every guard runs here, so a failure aborts before anything is written.

# COMMAND ----------

with run.step():
    silver = transform_ons(
        raw_df=raw,
        source_file=SOURCE_PATH,
        ingestion_ts=INGESTION_TS,
    )

    # One pass carries every measure below. last_measured is separate from last_month
    # because the panel is rectangular in date while Northern Ireland lags in value:
    # its trailing rows exist and carry a null index.
    by_area = (
        silver.groupBy("area_name")
        .agg(
            F.count(F.lit(1)).alias("rows"),
            F.max("date").alias("last_month"),
            F.max(
                F.when(F.col("price_index").isNotNull(), F.col("date"))
            ).alias("last_measured"),
        )
        .collect()
    )

    silver_rows = sum(row["rows"] for row in by_area)
    newest_month = max(row["last_month"] for row in by_area)

    run.measure("silver_rows", silver_rows)
    run.measure("geographies", len(by_area))
    # The filename labels a release; this is what the file holds. PIPR data lags its
    # release label by a month, so these differ by design.
    lag = run.freshness(newest_month, skip=SKIP_FRESHNESS)

print(f"{silver_rows:,} rows after typing")
print(f"{len(by_area):,} geographies")
print(f"filename vintage {source_file.name}, newest month in data {newest_month}")
print(f"{lag} days between that month and this run")

# COMMAND ----------

# MAGIC %md
# MAGIC ## Geography coverage
# MAGIC
# MAGIC Which geographies carry a published index in the newest month, against those
# MAGIC active in the last twelve. Every geography has a row in every month, so absence
# MAGIC here means the value is unpublished rather than the row missing. That is the
# MAGIC Northern Ireland lag: its rental areas and the Northern Ireland aggregate trail
# MAGIC the rest by about two months, and the UK figure only reaches the newest month
# MAGIC because ONS imputes them forward.
# MAGIC
# MAGIC The same twelve-month window as HPI, so the metric means the same thing on both
# MAGIC even though the panel shapes differ.

# COMMAND ----------

with run.step():
    # First month of the twelve ending at newest_month. Month arithmetic on a
    # zero-based ordinal: year * 12 + month - 1 keeps December in its own year, which
    # a year * 12 + month form does not.
    _ordinal = newest_month.year * 12 + newest_month.month - 1 - 11
    window_start = date(_ordinal // 12, _ordinal % 12 + 1, 1)

    active = [
        row
        for row in by_area
        if row["last_measured"] is not None and row["last_measured"] >= window_start
    ]
    absent = sorted(
        row["area_name"] for row in active if row["last_measured"] != newest_month
    )

    run.measure(
        "entities_in_newest_period",
        len(active) - len(absent),
        denominator=len(active),
    )
    if absent:
        run.measure("entities_absent_from_newest_period", absent)

print(f"{len(active) - len(absent):,} of {len(active):,} geographies publish an index "
      f"for {newest_month}")
print(f"unpublished ({len(absent)}): {absent or 'none'}")

# COMMAND ----------

# MAGIC %md
# MAGIC ## Create the Silver table and apply CHECK constraints
# MAGIC
# MAGIC The column list is generated from the transform module, so the DDL cannot drift
# MAGIC from the cast types. The table is created once (`IF NOT EXISTS`); constraints are
# MAGIC dropped and re-added each run so the notebook is idempotent. Single-row
# MAGIC invariants live here; multi-row invariants (unique grain, marker positions, the
# MAGIC unpublished months being the trailing ones) belong in the chispa test.
# MAGIC
# MAGIC No partitioning. The table is under 50,000 rows and is rewritten whole.

# COMMAND ----------

_ = spark.sql(  # noqa: F821
    f"""
    CREATE TABLE IF NOT EXISTS {TARGET_TABLE} (
    {silver_table_ddl()}
    )
    USING DELTA
    COMMENT 'ONS Price Index of Private Rents as a monthly geography panel, one row per (area_name, date) from January 2015. Keyed on area name because ONS publishes no code for the eight Northern Irish broad rental market areas. Geography is local authority district in England and Wales and broad rental market area in Scotland and Northern Ireland, so only the English and Welsh codes join to HPI and Doogal. City of London and Isles of Scilly are not published. Index base is January 2023 = 100 and the series is not seasonally adjusted. The latest UK months are part imputed while Northern Ireland lags; Great Britain covers the same months fully measured.'
    """
)

# COMMAND ----------

# MAGIC %sql
# MAGIC ALTER TABLE uk_property_intel.silver.ons_private_rents
# MAGIC   DROP CONSTRAINT IF EXISTS date_is_month_start;
# MAGIC ALTER TABLE uk_property_intel.silver.ons_private_rents
# MAGIC   ADD CONSTRAINT date_is_month_start
# MAGIC   CHECK (day(date) = 1);
# MAGIC
# MAGIC ALTER TABLE uk_property_intel.silver.ons_private_rents
# MAGIC   DROP CONSTRAINT IF EXISTS date_at_or_after_series_start;
# MAGIC ALTER TABLE uk_property_intel.silver.ons_private_rents
# MAGIC   ADD CONSTRAINT date_at_or_after_series_start
# MAGIC   CHECK (date >= DATE '2015-01-01');
# MAGIC
# MAGIC ALTER TABLE uk_property_intel.silver.ons_private_rents
# MAGIC   DROP CONSTRAINT IF EXISTS price_index_positive;
# MAGIC ALTER TABLE uk_property_intel.silver.ons_private_rents
# MAGIC   ADD CONSTRAINT price_index_positive
# MAGIC   CHECK (price_index IS NULL OR price_index > 0);
# MAGIC
# MAGIC ALTER TABLE uk_property_intel.silver.ons_private_rents
# MAGIC   DROP CONSTRAINT IF EXISTS rental_price_positive;
# MAGIC ALTER TABLE uk_property_intel.silver.ons_private_rents
# MAGIC   ADD CONSTRAINT rental_price_positive
# MAGIC   CHECK (rental_price IS NULL OR rental_price > 0);
# MAGIC
# MAGIC ALTER TABLE uk_property_intel.silver.ons_private_rents
# MAGIC   DROP CONSTRAINT IF EXISTS area_code_is_gss_or_absent;
# MAGIC ALTER TABLE uk_property_intel.silver.ons_private_rents
# MAGIC   ADD CONSTRAINT area_code_is_gss_or_absent
# MAGIC   CHECK (area_code IS NULL OR area_code RLIKE '^[A-Z][0-9]{8}$');

# COMMAND ----------

# MAGIC %md
# MAGIC ## Write (full refresh)
# MAGIC
# MAGIC `INSERT OVERWRITE` preserves the table schema, comment, and CHECK constraints;
# MAGIC only the data is replaced.

# COMMAND ----------

with run.step():
    # INSERT OVERWRITE matches on position, so a projection that drifts from the DDL
    # would load values into the wrong columns.
    assert tuple(silver.columns) == SILVER_COLUMNS, silver.columns

    silver.createOrReplaceTempView("_ons_silver_staging")
    spark.sql(  # noqa: F821
        f"INSERT OVERWRITE {TARGET_TABLE} TABLE _ons_silver_staging"
    )
    written = spark.table(TARGET_TABLE).count()  # noqa: F821

run.succeed(rows_written=written)
print(f"Wrote {written:,} rows to {TARGET_TABLE}")
print(f"run {run.run_id} recorded as succeeded")

# COMMAND ----------

# The read is complete, so the converted copy and the cached parse can go. Local disk is
# shared with the next notebook on this cluster.
_ = raw.unpersist()
os.remove(STAGE_PATH)
print(f"cleared {STAGE_PATH}")

# COMMAND ----------

# MAGIC %md
# MAGIC ## Verify

# COMMAND ----------

# MAGIC %sql
# MAGIC SELECT count(*) AS rows,
# MAGIC        count(DISTINCT area_name) AS geographies,
# MAGIC        count(DISTINCT area_code) AS coded_geographies,
# MAGIC        min(date) AS first_month,
# MAGIC        max(date) AS last_month
# MAGIC FROM uk_property_intel.silver.ons_private_rents

# COMMAND ----------

# MAGIC %md
# MAGIC The panel is rectangular: every geography carries every month, with none starting
# MAGIC late and none ending early. This returns one row. Two rows means a geography was
# MAGIC added or withdrawn mid-series, which is a boundary change rather than a fault, but
# MAGIC it changes what the grain guarantees.

# COMMAND ----------

# MAGIC %sql
# MAGIC SELECT months, count(*) AS geographies
# MAGIC FROM (
# MAGIC   SELECT area_name, count(*) AS months
# MAGIC   FROM uk_property_intel.silver.ons_private_rents
# MAGIC   GROUP BY area_name
# MAGIC )
# MAGIC GROUP BY months
# MAGIC ORDER BY months

# COMMAND ----------

# MAGIC %md
# MAGIC Geography inventory. E06 to E09 are English local authorities, W06 Welsh unitary
# MAGIC authorities, S33 Scottish rental market areas, E12 English regions, and the null
# MAGIC block is the eight uncoded Northern Irish rental areas. E09 is one short of 33 and
# MAGIC E06 one short of 63: City of London and Isles of Scilly are not published.

# COMMAND ----------

# MAGIC %sql
# MAGIC SELECT coalesce(left(area_code, 3), 'none (NI rental areas)') AS prefix,
# MAGIC        count(DISTINCT area_name) AS geographies
# MAGIC FROM uk_property_intel.silver.ons_private_rents
# MAGIC GROUP BY 1
# MAGIC ORDER BY prefix

# COMMAND ----------

# MAGIC %md
# MAGIC Join surface against HPI, which is what this table exists to sit beside. Only the
# MAGIC English and Welsh local authorities match: Scotland reports by rental market area
# MAGIC rather than council area, so S33 finds nothing, and the Northern Irish areas carry
# MAGIC no code to join on at all.

# COMMAND ----------

# MAGIC %sql
# MAGIC SELECT left(o.area_code, 3) AS prefix,
# MAGIC        count(*) AS ons_areas,
# MAGIC        count_if(h.area_code IS NOT NULL) AS matched_in_hpi
# MAGIC FROM (
# MAGIC   SELECT DISTINCT area_code
# MAGIC   FROM uk_property_intel.silver.ons_private_rents
# MAGIC   WHERE area_code IS NOT NULL
# MAGIC ) o
# MAGIC LEFT JOIN (
# MAGIC   SELECT DISTINCT area_code FROM uk_property_intel.silver.hpi
# MAGIC ) h ON h.area_code = o.area_code
# MAGIC GROUP BY 1
# MAGIC ORDER BY prefix

# COMMAND ----------

# MAGIC %md
# MAGIC The Northern Ireland lag. Its last measured month sits behind every other nation,
# MAGIC and the UK figure runs to the same month as the rest because ONS imputes Northern
# MAGIC Ireland forward to produce it. Great Britain excludes Northern Ireland and is
# MAGIC unaffected.

# COMMAND ----------

# MAGIC %sql
# MAGIC SELECT area_name,
# MAGIC        max(date) AS last_month,
# MAGIC        max(CASE WHEN price_index IS NOT NULL THEN date END) AS last_measured,
# MAGIC        count_if(price_index IS NULL) AS unpublished_months
# MAGIC FROM uk_property_intel.silver.ons_private_rents
# MAGIC WHERE area_code IN ('K02000001', 'K03000001', 'E92000001', 'W92000004',
# MAGIC                     'S92000003', 'N92000002')
# MAGIC GROUP BY area_name
# MAGIC ORDER BY last_measured, area_name

# COMMAND ----------

# MAGIC %md
# MAGIC Structural nulls. Annual change is absent for the first twelve months because
# MAGIC there is no prior year to compare against, and monthly change for the first month
# MAGIC alone. Outside those, a null here means Northern Ireland has not been published.

# COMMAND ----------

# MAGIC %sql
# MAGIC WITH bounds AS (
# MAGIC   SELECT max(date) AS last_month FROM uk_property_intel.silver.ons_private_rents
# MAGIC )
# MAGIC SELECT r.date,
# MAGIC        count_if(r.pct_change_1m IS NULL) AS no_monthly_change,
# MAGIC        count_if(r.pct_change_12m IS NULL) AS no_annual_change,
# MAGIC        count(*) AS geographies
# MAGIC FROM uk_property_intel.silver.ons_private_rents r, bounds b
# MAGIC WHERE r.date < DATE '2016-02-01'
# MAGIC    OR r.date >= add_months(b.last_month, -2)
# MAGIC GROUP BY r.date
# MAGIC ORDER BY r.date

# COMMAND ----------

# MAGIC %sql
# MAGIC SELECT date, area_name, price_index, pct_change_1m, pct_change_12m, rental_price,
# MAGIC        one_bed_rental_price, two_bed_rental_price, flat_maisonette_rental_price
# MAGIC FROM uk_property_intel.silver.ons_private_rents
# MAGIC WHERE area_code = 'K02000001'
# MAGIC ORDER BY date DESC
# MAGIC LIMIT 12

# COMMAND ----------

# MAGIC %sql
# MAGIC SELECT date, area_name, region_or_country_name, price_index, rental_price
# MAGIC FROM uk_property_intel.silver.ons_private_rents
# MAGIC WHERE area_code IS NULL
# MAGIC   AND date = (SELECT max(date) FROM uk_property_intel.silver.ons_private_rents)
# MAGIC ORDER BY area_name

# COMMAND ----------

# MAGIC %md
# MAGIC ## What this run recorded

# COMMAND ----------

# MAGIC %sql
# MAGIC SELECT r.status, r.started_ts, r.ended_ts, r.rows_written, r.error_type,
# MAGIC        m.metric, m.value_numeric, m.value_text, m.value_date, m.denominator
# MAGIC FROM uk_property_intel.quality.pipeline_run r
# MAGIC LEFT JOIN uk_property_intel.quality.pipeline_metric m USING (run_id)
# MAGIC WHERE r.source = 'ons'
# MAGIC ORDER BY r.started_ts DESC, m.metric
